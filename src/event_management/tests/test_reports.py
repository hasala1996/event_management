from io import BytesIO
from rest_framework import status
import openpyxl
from django.utils.timezone import now
from core.utils.test_setup import TestSetup
from faker import Faker
from event_management.models import Category, Event

faker = Faker()


class TestEventReportEndpoint(TestSetup):
    """
    Test case for the event report generation endpoint.
    """

    def setup_method(self, method):
        """
        Setup common elements for each test.
        """
        Category.objects.all().delete()
        Event.objects.all().delete()
        self.category1 = self.create_category(
            name=faker.name(), description="Tech events"
        )
        self.category2 = self.create_category(
            name=faker.name(), description="Sports events"
        )

        self.event1 = self.create_event(
            name="Tech Conference",
            description="A tech conference",
            date=now(),
            location="San Francisco",
            category=self.category1,
            total_slots=100,
            is_featured=True,
        )
        self.event2 = self.create_event(
            name="Sports Meetup",
            description="A sports meetup",
            date=now(),
            location="Los Angeles",
            category=self.category2,
            total_slots=50,
            is_featured=False,
        )

        self.url = "/v1/api/event_management/event/generate-report/"

    def test_generate_report_with_category_filter(self):
        """
        Test report generation with category filter. It should only include events in the specified category.
        """
        response = self.client.get(f"{self.url}?category_id={self.category1.id}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        expected_headers = [
            "Event Name",
            "Description",
            "Date",
            "Location",
            "Category",
            "Is Featured",
        ]
        self.assertEqual(headers, expected_headers)

        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 1)

        expected_row = (
            self.event1.name,
            self.event1.description,
            self.event1.date.strftime("%Y-%m-%d %H:%M:%S"),
            self.event1.location,
            self.event1.category.name,
            ("Yes" if self.event1.is_featured else "No"),
        )
        self.assertIn(expected_row, data_rows)

    def test_generate_report_with_date_filters(self):
        """
        Test report generation with date filters. It should only include events within the specified date range.
        """
        start_date = self.event1.date.strftime("%Y-%m-%d")
        response = self.client.get(f"{self.url}?start_date={start_date}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active

        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 2)

    def test_invalid_filters(self):
        """
        Test report generation with invalid filters. It should return a 400 error.
        """
        response = self.client.get(self.url, {"start_date": "invalid-date"})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

        response_data = response.json()

        self.assertIn("start_date", response_data)

        expected_error_message = [
            "Date has wrong format. Use one of these formats instead: YYYY-MM-DD."
        ]
        self.assertEqual(response_data["start_date"], expected_error_message)
