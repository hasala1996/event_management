from rest_framework import serializers
from openpyxl import Workbook
from io import BytesIO
from event_management.models import Event
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


class EventReportSerializer(serializers.Serializer):
    """
    Serializer for validating filters and generating event reports.
    """

    category_id = serializers.UUIDField(
        required=False, help_text="Filter by category ID"
    )
    start_date = serializers.DateField(
        required=False, help_text="Filter by start date (YYYY-MM-DD)"
    )
    end_date = serializers.DateField(
        required=False, help_text="Filter by end date (YYYY-MM-DD)"
    )

    def generate_report(self):
        """
        Generate an Excel report based on the validated filters.

        Returns:
            BytesIO: The generated Excel file as a binary stream.
        """
        category_id = self.validated_data.get("category_id")
        start_date = self.validated_data.get("start_date")
        end_date = self.validated_data.get("end_date")

        events = Event.objects.all()
        if category_id:
            events = events.filter(category_id=category_id)
        if start_date:
            events = events.filter(date__gte=start_date)
        if end_date:
            events = events.filter(date__lte=end_date)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Event Report"

        headers = [
            "Event Name",
            "Description",
            "Date",
            "Location",
            "Category",
            "Is Featured",
        ]
        sheet.append(headers)
        header_fill = PatternFill(
            start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
        )
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for event in events:
            sheet.append(
                [
                    event.name,
                    event.description,
                    event.date.strftime("%Y-%m-%d %H:%M:%S"),
                    event.location,
                    event.category.name,
                    "Yes" if event.is_featured else "No",
                ]
            )

        table_ref = f"A1:F{sheet.max_row}"
        table = Table(displayName="EventTable", ref=table_ref)
        sheet.add_table(table)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
